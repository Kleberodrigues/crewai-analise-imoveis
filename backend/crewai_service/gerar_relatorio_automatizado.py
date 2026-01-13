"""
Gerador de Relatorio Automatizado - Analise de Imovel de Leilao
Gera relatorio completo em Excel com pesquisa de mercado real
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from tools.data_tools import pesquisar_mercado, comparar_imovel_mercado


def formatar_moeda(valor):
    """Formata valor como moeda brasileira"""
    if valor is None:
        return "R$ 0,00"
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def gerar_relatorio_excel(imovel: dict, output_path: str = None) -> str:
    """
    Gera relatorio completo em Excel para um imovel de leilao.

    Args:
        imovel: Dicionario com dados do imovel
        output_path: Caminho do arquivo de saida (opcional)

    Returns:
        Caminho do arquivo gerado
    """

    # 1. Pesquisa de mercado
    print("[1/5] Pesquisando mercado...")
    mercado = pesquisar_mercado(
        bairro=imovel.get("bairro", "").lower(),
        cidade="sao-paulo",
        uf="sp",
        tipo=imovel.get("tipo_imovel", "apartamento").lower(),
        quartos=imovel.get("quartos", 2),
        area_m2=imovel.get("area_privativa", 50),
        fontes=['vivareal', 'olx']
    )

    # 2. Comparacao com mercado
    print("[2/5] Comparando com mercado...")
    comparacao = comparar_imovel_mercado(imovel, mercado)

    # 3. Calculos de custos
    print("[3/5] Calculando custos...")

    preco_compra = imovel.get("preco", 0)
    area = imovel.get("area_privativa", 50)
    cidade = imovel.get("cidade", "SAO PAULO")

    # Custos de aquisicao
    itbi = preco_compra * 0.03  # 3% SP
    cartorio = 2500
    despachante = 800
    reforma = area * 500  # R$ 500/m2
    total_aquisicao = preco_compra + itbi + cartorio + despachante + reforma

    # Valor de venda
    valor_mercado = mercado.get("valor_estimado", {}).get("valor", area * 5000)
    valor_venda = valor_mercado * 0.95  # -5% para venda rapida

    # Custos de venda
    comissao = valor_venda * 0.06  # 6%
    lucro_antes_ir = valor_venda - total_aquisicao - comissao
    ir_ganho_capital = max(0, lucro_antes_ir * 0.15)  # 15%
    total_venda = comissao + ir_ganho_capital

    # Resultado
    lucro_liquido = valor_venda - total_aquisicao - total_venda
    roi = (lucro_liquido / total_aquisicao) * 100 if total_aquisicao > 0 else 0
    margem = (lucro_liquido / valor_venda) * 100 if valor_venda > 0 else 0

    # Custos mensais (holding)
    condominio = 400
    iptu_mensal = 158
    agua_luz = 100
    total_mensal = condominio + iptu_mensal + agua_luz

    # 4. Criar Excel
    print("[4/5] Gerando Excel...")

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    money_font = Font(bold=True, color="006400", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === ABA 1: RESUMO ===
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    resumo_data = [
        ["ANALISE DE IMOVEL DE LEILAO", ""],
        ["Data da Analise", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["", ""],
        ["DADOS DO IMOVEL", ""],
        ["Endereco", imovel.get("endereco", "N/I")],
        ["Bairro", imovel.get("bairro", "N/I")],
        ["Cidade/UF", f"{imovel.get('cidade', 'N/I')}/{imovel.get('uf', 'SP')}"],
        ["Tipo", imovel.get("tipo_imovel", "Apartamento")],
        ["Area Privativa", f"{area:.2f} m2"],
        ["Quartos", imovel.get("quartos", 2)],
        ["Vagas", imovel.get("vagas", 1)],
        ["", ""],
        ["VALORES", ""],
        ["Preco de Venda (Caixa)", formatar_moeda(preco_compra)],
        ["Valor de Avaliacao", formatar_moeda(imovel.get("valor_avaliacao", 0))],
        ["Desconto sobre Avaliacao", f"{imovel.get('desconto', 0):.1f}%"],
        ["", ""],
        ["PESQUISA DE MERCADO", ""],
        ["Fonte", mercado.get("fonte", "N/A")],
        ["Preco Mercado Medio", formatar_moeda(mercado.get("precos", {}).get("medio", 0))],
        ["Preco por m2", formatar_moeda(mercado.get("preco_m2", {}).get("medio", 0)) + "/m2"],
        ["Imoveis Comparaveis", mercado.get("total_encontrados", 0)],
        ["Desconto vs Mercado", f"{comparacao.get('desconto_vs_mercado', 0):.1f}%"],
        ["Classificacao", comparacao.get("classificacao_oportunidade", "N/A")],
        ["", ""],
        ["RESULTADO FINANCEIRO", ""],
        ["Total Aquisicao", formatar_moeda(total_aquisicao)],
        ["Valor de Venda Estimado", formatar_moeda(valor_venda)],
        ["Lucro Liquido", formatar_moeda(lucro_liquido)],
        ["ROI", f"{roi:.1f}%"],
        ["Margem Liquida", f"{margem:.1f}%"],
        ["", ""],
        ["RECOMENDACAO", comparacao.get("classificacao_oportunidade", "")],
    ]

    for row_idx, (label, value) in enumerate(resumo_data, 1):
        ws_resumo.cell(row=row_idx, column=1, value=label)
        ws_resumo.cell(row=row_idx, column=2, value=value)

        # Formatacao de headers
        if label in ["ANALISE DE IMOVEL DE LEILAO", "DADOS DO IMOVEL", "VALORES",
                     "PESQUISA DE MERCADO", "RESULTADO FINANCEIRO", "RECOMENDACAO"]:
            ws_resumo.cell(row=row_idx, column=1).font = header_font
            ws_resumo.cell(row=row_idx, column=1).fill = header_fill
            ws_resumo.cell(row=row_idx, column=2).fill = header_fill

    ws_resumo.column_dimensions['A'].width = 30
    ws_resumo.column_dimensions['B'].width = 40

    # === ABA 2: CUSTOS DE AQUISICAO ===
    ws_custos = wb.create_sheet("Custos Aquisicao")

    custos_data = [
        ["Item", "Valor", "Observacao"],
        ["Preco de Compra", preco_compra, "Valor do imovel na Caixa"],
        ["ITBI (3%)", itbi, "Imposto de transmissao - SP"],
        ["Cartorio", cartorio, "Registro e escritura"],
        ["Despachante", despachante, "Honorarios"],
        ["Reforma Estimada", reforma, f"R$ 500/m2 x {area:.2f}m2"],
        ["TOTAL", total_aquisicao, ""],
    ]

    for row_idx, row_data in enumerate(custos_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_custos.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border

    ws_custos.column_dimensions['A'].width = 25
    ws_custos.column_dimensions['B'].width = 20
    ws_custos.column_dimensions['C'].width = 35

    # === ABA 3: CUSTOS DE VENDA ===
    ws_venda = wb.create_sheet("Custos Venda")

    venda_data = [
        ["Item", "Valor", "Observacao"],
        ["Valor de Venda", valor_venda, "95% do valor de mercado"],
        ["Comissao Corretor (6%)", comissao, "Corretagem padrao"],
        ["IR Ganho de Capital (15%)", ir_ganho_capital, "Sobre lucro"],
        ["Total Custos Venda", total_venda, ""],
        ["", "", ""],
        ["LUCRO LIQUIDO", lucro_liquido, "Venda - Aquisicao - Custos"],
        ["ROI", f"{roi:.1f}%", "Lucro / Investimento"],
    ]

    for row_idx, row_data in enumerate(venda_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_venda.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border

    ws_venda.column_dimensions['A'].width = 30
    ws_venda.column_dimensions['B'].width = 20
    ws_venda.column_dimensions['C'].width = 35

    # === ABA 4: CENARIOS DE VENDA ===
    ws_cenarios = wb.create_sheet("Cenarios")

    cenarios = [
        ["Cenario", "Valor Venda", "Lucro Liquido", "ROI"],
        ["Pessimista (-10%)", valor_mercado * 0.90, 0, 0],
        ["Conservador (-5%)", valor_mercado * 0.95, 0, 0],
        ["Moderado (mercado)", valor_mercado, 0, 0],
        ["Otimista (+5%)", valor_mercado * 1.05, 0, 0],
    ]

    # Calcula lucros para cada cenario
    for i in range(1, len(cenarios)):
        vv = cenarios[i][1]
        com = vv * 0.06
        lucro_bruto = vv - total_aquisicao - com
        ir = max(0, lucro_bruto * 0.15)
        ll = lucro_bruto - ir
        roi_cenario = (ll / total_aquisicao) * 100 if total_aquisicao > 0 else 0
        cenarios[i][2] = ll
        cenarios[i][3] = f"{roi_cenario:.1f}%"

    for row_idx, row_data in enumerate(cenarios, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_cenarios.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border

    ws_cenarios.column_dimensions['A'].width = 25
    ws_cenarios.column_dimensions['B'].width = 20
    ws_cenarios.column_dimensions['C'].width = 20
    ws_cenarios.column_dimensions['D'].width = 15

    # === ABA 5: CUSTOS MENSAIS ===
    ws_holding = wb.create_sheet("Custos Mensais")

    holding_data = [
        ["Item", "Valor Mensal", "3 Meses", "6 Meses", "9 Meses"],
        ["Condominio", condominio, condominio*3, condominio*6, condominio*9],
        ["IPTU", iptu_mensal, iptu_mensal*3, iptu_mensal*6, iptu_mensal*9],
        ["Agua/Luz", agua_luz, agua_luz*3, agua_luz*6, agua_luz*9],
        ["TOTAL", total_mensal, total_mensal*3, total_mensal*6, total_mensal*9],
    ]

    for row_idx, row_data in enumerate(holding_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_holding.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border

    ws_holding.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws_holding.column_dimensions[col].width = 15

    # 5. Salvar arquivo
    print("[5/5] Salvando arquivo...")

    if output_path is None:
        bairro_safe = imovel.get("bairro", "imovel").replace(" ", "_").lower()
        output_path = f"relatorio_{bairro_safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    wb.save(output_path)
    print(f"\nRelatorio salvo: {output_path}")

    return output_path


# Dados do imovel Estacao Primavera
IMOVEL_ESTACAO_PRIMAVERA = {
    "id_imovel": "estacao_primavera_01",
    "endereco": "Rua Eduardo Rizk, 200 - Apto 23 Bloco A - Ed. Residencial Estacao Primavera",
    "bairro": "Guaianazes",
    "cidade": "SAO PAULO",
    "uf": "SP",
    "tipo_imovel": "Apartamento",
    "area_privativa": 38.72,
    "quartos": 2,
    "vagas": 1,
    "preco": 78000.00,
    "valor_avaliacao": 178000.00,
    "desconto": 56.2,
    "link": "https://venda-imoveis.caixa.gov.br/sistema/detalhe-imovel.asp"
}


if __name__ == "__main__":
    print("=" * 60)
    print("GERADOR DE RELATORIO AUTOMATIZADO")
    print("Imovel: Estacao Primavera - Guaianazes")
    print("=" * 60)

    output_file = gerar_relatorio_excel(
        IMOVEL_ESTACAO_PRIMAVERA,
        "relatorio_estacao_primavera_automatizado.xlsx"
    )

    print("\n" + "=" * 60)
    print("RELATORIO GERADO COM SUCESSO!")
    print(f"Arquivo: {output_file}")
    print("=" * 60)
